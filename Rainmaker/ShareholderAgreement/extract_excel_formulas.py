import openpyxl

# Path to your Excel file
excel_path = "CapGov_Templatev2.xlsx"

# Open the workbook
wb = openpyxl.load_workbook(excel_path, data_only=False)

output_lines = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    output_lines.append(f"# Sheet: {sheet}")
    for row in ws.iter_rows():
        row_values = []
        for cell in row:
            if cell.value is not None:
                if cell.data_type == 'f':
                    row_values.append(f"FORMULA: {cell.coordinate} = {cell.value}")
                else:
                    row_values.append(f"{cell.coordinate}: {cell.value}")
        if row_values:
            output_lines.append(" | ".join(row_values))
    output_lines.append("")

# Write output to a text file
with open("CapGov_Templatev2_formulas_and_values.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(output_lines))

print("Extraction complete. See CapGov_Templatev2_formulas_and_values.txt for details.")
