import openpyxl
from openpyxl.styles import Font
from datetime import datetime

# --- Step 1: Create Expanded Main Ledger Sheet ---
wb = openpyxl.Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

ledger = wb.create_sheet(title="Main Ledger")
ledger_headers = [
    "Date", "Transaction Type", "From", "To", "Member", "Share Class", "Shares", "Price/Share (USD)", "Total Value (USD)", "Vesting/Unlock Date", "Conversion Details", "Notes"
]
for col, header in enumerate(ledger_headers, 1):
    cell = ledger.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

sample_data = [
    ["2025-05-14", "Issuance", "Company", "Founder", "Founder", "Founder", 5100000, 1.00, 5100000.00, "2025-12-31", "", "Initial allocation"],
    ["2025-05-14", "Issuance", "Company", "Member_1", "Member_1", "Class A", 490000, 1.00, 490000.00, "2026-12-31", "", "Initial allocation"],
    ["2025-05-14", "Issuance", "Company", "Member_2", "Member_2", "Class A", 490000, 1.00, 490000.00, "2026-12-31", "", "Initial allocation"],
    ["2025-05-14", "Issuance", "Company", "Member_3", "Member_3", "Class A", 490000, 1.00, 490000.00, "2026-12-31", "", "Initial allocation"],
    ["2025-05-14", "Issuance", "Company", "Investor_1", "Investor_1", "Class B", 100000, 2.00, 200000.00, "", "", "First funding round"],
    ["2025-05-14", "Issuance", "Company", "Investor_2", "Investor_2", "Class C", 50000, 3.00, 150000.00, "", "", "First funding round"],
    ["2026-01-01", "Vesting", "Founder", "Founder", "Founder", "Founder", 510000, 1.00, 510000.00, "2026-12-31", "", "10% unlock"],
    ["2026-06-01", "Transfer", "Founder", "Investor_1", "Investor_1", "Class B", 10000, 2.00, 20000.00, "", "Converted from Founder Shares", "Founder sells to Investor_1"],
    ["2026-07-01", "Conversion", "Member_1", "Member_1", "Member_1", "Class B", 10000, 2.00, 20000.00, "", "Class A to Class B", "Conversion event"],
]
for i, row in enumerate(sample_data, 2):
    for j, val in enumerate(row, 1):
        ledger.cell(row=i, column=j, value=val)

# --- Step 2: Create Founder Dashboard (Ledger-Driven) ---
founder_sheet = wb.create_sheet(title="Founder Dashboard")
headers = ["Share Class", "Total Shares Held", "Vested Shares", "Unvested Shares", "Voting Power", "Dividend Multiple", "Transaction History"]
for col, header in enumerate(headers, 1):
    founder_sheet.cell(row=1, column=col, value=header).font = Font(bold=True)

share_classes = ["Founder", "Class A", "Class B", "Class C"]
voting_power = {"Founder": 10, "Class A": 6, "Class B": 1, "Class C": 0}
dividend_mult = {"Founder": 1.0, "Class A": 1.0, "Class B": 1.0, "Class C": 2.0}

for i, sc in enumerate(share_classes, 2):
    founder_sheet.cell(row=i, column=1, value=sc)
    # Total Shares Held
    founder_sheet.cell(row=i, column=2, value=f"=SUMIFS('Main Ledger'!G:G, 'Main Ledger'!E:E, \"Founder\", 'Main Ledger'!F:F, \"{sc}\")")
    # Vested Shares (if vesting date <= today)
    founder_sheet.cell(row=i, column=3, value=f"=SUMIFS('Main Ledger'!G:G, 'Main Ledger'!E:E, \"Founder\", 'Main Ledger'!F:F, \"{sc}\", 'Main Ledger'!J:J, \"<=\"&TODAY())")
    # Unvested Shares (if vesting date > today)
    founder_sheet.cell(row=i, column=4, value=f"=SUMIFS('Main Ledger'!G:G, 'Main Ledger'!E:E, \"Founder\", 'Main Ledger'!F:F, \"{sc}\", 'Main Ledger'!J:J, \">\"&TODAY())")
    # Voting Power
    founder_sheet.cell(row=i, column=5, value=f"=B{i}*{voting_power[sc]}")
    # Dividend Multiple
    founder_sheet.cell(row=i, column=6, value=dividend_mult[sc])
    # Transaction History (user filter note)
    founder_sheet.cell(row=i, column=7, value="See Main Ledger for full transaction history.")

# --- Step 3: Save Workbook ---
wb.save("Rainmaker_Shareholder_Model.xlsx")
print("Expanded Main Ledger and Founder Dashboard created in 'Rainmaker_Shareholder_Model.xlsx'.")
